import csv
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
INPUT_DIR = ROOT / "Input files" / "Locations"
OUTPUT_FILE = ROOT / "Output" / "1_Initial" / "Location_Details_Preliminary_Delta.csv"


def _to_float(value: object | None) -> float | None:
    """Convert values to float and gracefully handle blanks or invalid strings."""
    if value is None:
        return None

    text = str(value).strip()
    if text == "":
        return None

    try:
        return float(text)
    except ValueError:
        return None


def prepare_location_delta() -> Path:
    source_file = INPUT_DIR / "Location details.xlsx"
    if not source_file.exists():
        raise FileNotFoundError("Missing source file: Input files/Locations/Location details.xlsx")

    workbook = load_workbook(source_file, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Source worksheet is empty.")

    header_cells = [str(value).strip() if value is not None else "" for value in rows[0]]
    indexed_headers = [
        (index, header)
        for index, header in enumerate(header_cells)
        if header != ""
    ]
    headers = [header for _, header in indexed_headers]

    required_columns = ["Location", "Location height", "Item height"]
    missing_columns = [column for column in required_columns if column not in headers]
    if missing_columns:
        raise KeyError(f"Missing required columns: {', '.join(missing_columns)}")

    output_fields = list(headers)
    if "Delta" not in output_fields:
        output_fields.append("Delta")

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with OUTPUT_FILE.open("w", newline="", encoding="utf-8") as target:
        writer = csv.DictWriter(target, fieldnames=output_fields)
        writer.writeheader()

        for values in rows[1:]:
            row = {}
            for index, header in indexed_headers:
                row[header] = values[index] if index < len(values) else None

            location_height = _to_float(row.get("Location height"))
            item_height = _to_float(row.get("Item height"))

            if location_height is None or item_height is None:
                row["Delta"] = ""
            else:
                row["Delta"] = f"{location_height - item_height:g}"

            writer.writerow(row)

    return OUTPUT_FILE


if __name__ == "__main__":
    output_path = prepare_location_delta()
    print(f"Preliminary delta preparation complete. Output written to: {output_path}")
