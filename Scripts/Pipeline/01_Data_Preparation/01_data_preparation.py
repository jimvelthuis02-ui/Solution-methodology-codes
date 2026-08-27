import csv
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[3]
INPUT_DIR = ROOT / "Input files" / "Locations"
PREPARED_OUTPUT_FILE = ROOT / "Output" / "01_Data_Preparation" / "Location_Details_Prepared.csv"
BEAM_OUTPUT_DIR = ROOT / "Output" / "01_Data_Preparation"
CURRENT_SPACE_UTILIZATION_FILE = BEAM_OUTPUT_DIR / "Current_Layout_Space_Utilization.csv"

BEAM_HEIGHT_CM = 16.0
ROW_ORDER_PATTERN = re.compile(r"^(\d+)([A-Za-z]?)$")

# Explicit geometry overrides confirmed from warehouse layout notes.
FORCED_BEAM_POINTS: set[tuple[str, int, str]] = {
    ("F", 0, "05"),
    ("F", 1, "05"),
    ("F", 2, "05"),
    ("F", 3, "05"),
    ("F", 4, "06"),
    ("F", 5, "06"),
    ("F", 6, "06"),
    ("K", 7, "2a"),
    ("K", 7, "2b"),
}

DISABLED_BEAM_POINTS: set[tuple[str, int, str]] = {
    ("K", 7, "02"),
}

BEAM_ALIASES: dict[tuple[str, int, str], tuple[str, int, str]] = {
    ("K", 7, "02"): ("K", 7, "2b"),
}


def _to_float_optional(value: object | None) -> float | None:
    if value is None:
        return None
    text = str(value).strip()
    if text == "":
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _to_float_default(value: object, default: float = 0.0) -> float:
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return default


def _to_int(value: object, default: int = -1) -> int:
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return default


def _row_sort_key(row_label: object) -> tuple[int, str]:
    text = str(row_label).strip()
    match = ROW_ORDER_PATTERN.match(text)
    if match is None:
        return (10**9, text.lower())
    return (int(match.group(1)), match.group(2).lower())


def _location_sort_key(row: dict[str, str]) -> tuple[int, str, str]:
    return (*_row_sort_key(row.get("Row")), str(row.get("Location", "")).strip())


def calculate_current_layout_space_utilization() -> Path:
    """Calculate occupancy and space utilization for each recorded status snapshot."""
    rows = _read_prepared_rows()
    status_columns = list(
        dict.fromkeys(
            status_column
            for row in rows
            for status_column in row
            if str(status_column).strip().startswith("Status ")
        )
    )
    status_columns.sort(key=lambda status: (status != "Status Initial", status))
    summary_rows: list[dict[str, str]] = []

    for status_column in status_columns:
        total_locations = 0
        occupied_locations = 0
        empty_locations = 0
        occupied_space_cm = 0.0
        empty_space_cm = 0.0

        for row in rows:
            status = str(row.get(status_column, "")).strip().lower()
            slot_size_cm = _to_float_optional(row.get("Location height"))
            if status not in {"occupied", "empty"} or slot_size_cm is None or slot_size_cm < 0.0:
                continue

            total_locations += 1
            if status == "occupied":
                occupied_locations += 1
                occupied_space_cm += slot_size_cm
            else:
                empty_locations += 1
                empty_space_cm += slot_size_cm

        total_space_cm = occupied_space_cm + empty_space_cm
        summary_rows.append(
            {
                "Status_Measurement": status_column,
                "Total_Locations": str(total_locations),
                "Occupied_Locations": str(occupied_locations),
                "Empty_Locations": str(empty_locations),
                "Occupancy_Rate": f"{(occupied_locations / total_locations) if total_locations else 0.0:.6f}",
                "Occupied_Slot_Space_m3": f"{occupied_space_cm / 100.0:.6f}",
                "Empty_Slot_Space_m3": f"{empty_space_cm / 100.0:.6f}",
                "Total_Slot_Space_m3": f"{total_space_cm / 100.0:.6f}",
                "Space_Utilization_Pct": f"{(occupied_space_cm / total_space_cm * 100.0) if total_space_cm else 0.0:.4f}",
            }
        )

    with CURRENT_SPACE_UTILIZATION_FILE.open("w", newline="", encoding="utf-8") as target:
        fieldnames = [
            "Status_Measurement",
            "Total_Locations",
            "Occupied_Locations",
            "Empty_Locations",
            "Occupancy_Rate",
            "Occupied_Slot_Space_m3",
            "Empty_Slot_Space_m3",
            "Total_Slot_Space_m3",
            "Space_Utilization_Pct",
        ]
        writer = csv.DictWriter(target, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(summary_rows)

    return CURRENT_SPACE_UTILIZATION_FILE


def _segment_bounds(max_col: int) -> list[tuple[int, int]]:
    # Segment columns according to the warehouse beam structure:
    # first segment spans 0-3, remaining segments span groups of 3.
    if max_col < 0:
        return []

    bounds: list[tuple[int, int]] = []
    start = 0

    first_end = min(max_col, 3)
    bounds.append((start, first_end))
    start = first_end + 1

    while start <= max_col:
        end = min(max_col, start + 2)
        bounds.append((start, end))
        start = end + 1

    return bounds


def _is_row_without_beam_support(row_label: str) -> bool:
    # Row 1 / 1a are floor-level positions and are treated as non-beam support.
    match = ROW_ORDER_PATTERN.match(row_label.strip())
    if match is None:
        return False
    row_num = int(match.group(1))
    suffix = match.group(2).lower()
    return row_num == 1 and suffix in ("", "a")


def prepare_location_data() -> Path:
    """Read raw location workbook and write a cleaned CSV with Delta height."""
    source_file = INPUT_DIR / "Location details.xlsx"
    if not source_file.exists():
        raise FileNotFoundError("Missing source file: Input files/Locations/Location details.xlsx")

    workbook = load_workbook(source_file, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Source worksheet is empty.")

    header_cells = [str(value).strip() if value is not None else "" for value in rows[0]]
    indexed_headers = [(index, header) for index, header in enumerate(header_cells) if header]
    headers = [header for _, header in indexed_headers]

    required_columns = ["Location", "Location height", "Item height"]
    missing = [column for column in required_columns if column not in headers]
    if missing:
        raise KeyError(f"Missing required columns: {', '.join(missing)}")

    output_fields = list(headers)
    if "Delta" not in output_fields:
        output_fields.append("Delta")

    PREPARED_OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with PREPARED_OUTPUT_FILE.open("w", newline="", encoding="utf-8") as target:
        writer = csv.DictWriter(target, fieldnames=output_fields)
        writer.writeheader()

        for values in rows[1:]:
            row: dict[str, object | None] = {}
            for index, header in indexed_headers:
                row[header] = values[index] if index < len(values) else None

            location_height = _to_float_optional(row.get("Location height"))
            item_height = _to_float_optional(row.get("Item height"))
            row["Delta"] = "" if location_height is None or item_height is None else f"{location_height - item_height:g}"
            writer.writerow(row)

    return PREPARED_OUTPUT_FILE


def _read_prepared_rows() -> list[dict[str, str]]:
    """Load the prepared location CSV emitted by `prepare_location_data`."""
    if not PREPARED_OUTPUT_FILE.exists():
        raise FileNotFoundError(f"Missing input file: {PREPARED_OUTPUT_FILE}")
    with PREPARED_OUTPUT_FILE.open("r", newline="", encoding="utf-8-sig") as source:
        reader = csv.DictReader(source)
        if reader.fieldnames is None:
            raise ValueError("Input CSV has no header row.")
        return list(reader)


def build_beam_grid_map() -> Path:
    """Build beam segments, location-to-beam mapping, and beam height coordinates."""
    rows = _read_prepared_rows()
    BEAM_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    grouped: dict[tuple[str, int], list[dict[str, str]]] = defaultdict(list)
    rack_columns: dict[str, set[int]] = defaultdict(set)
    row_index: dict[tuple[str, int, str], dict[str, str]] = {}

    # Index rows by (rack, column, row) to support fast beam lookups.
    for row in rows:
        rack = str(row.get("Rack", "")).strip()
        column_num = _to_int(row.get("Column"))
        if rack and column_num >= 0:
            grouped[(rack, column_num)].append(row)
            rack_columns[rack].add(column_num)
            row_label = str(row.get("Row", "")).strip()
            row_index[(rack, column_num, row_label)] = row

    beam_segments: list[dict[str, str]] = []
    location_map: list[dict[str, str]] = []
    beam_for_location: dict[tuple[str, int, str], dict[str, str]] = {}

    # Discover beam segments per rack and row, applying forced/disabled overrides.
    for rack in sorted(rack_columns):
        max_col = max(rack_columns[rack]) if rack_columns[rack] else -1
        segments = _segment_bounds(max_col)

        rack_row_labels = {
            str(row.get("Row", "")).strip()
            for (current_rack, _), rack_rows in grouped.items()
            if current_rack == rack
            for row in rack_rows
        }
        rack_row_labels.update(row_label for current_rack, _, row_label in FORCED_BEAM_POINTS if current_rack == rack)

        for row_label in sorted(rack_row_labels, key=_row_sort_key):
            if _is_row_without_beam_support(row_label):
                continue

            for seg_start, seg_end in segments:
                supported_cols: list[int] = []
                for col_num in range(seg_start, seg_end + 1):
                    key = (rack, col_num, row_label)
                    source_row = row_index.get(key)
                    beam_supported = False
                    if source_row is not None:
                        beam_text = str(source_row.get("Beam column count", "")).strip()
                        beam_supported = beam_text != ""
                    if key in FORCED_BEAM_POINTS:
                        beam_supported = True
                    if key in DISABLED_BEAM_POINTS:
                        beam_supported = False
                    if beam_supported:
                        supported_cols.append(col_num)

                if not supported_cols:
                    continue

                # Emit one beam object per contiguous supported run so span labels
                # reflect the real physical span (for example 2-column instead of
                # full 3-column when blocked by layout/split constraints).
                contiguous_runs: list[tuple[int, int]] = []
                run_start = supported_cols[0]
                run_end = supported_cols[0]
                for col_num in supported_cols[1:]:
                    if col_num == run_end + 1:
                        run_end = col_num
                        continue
                    contiguous_runs.append((run_start, run_end))
                    run_start = col_num
                    run_end = col_num
                contiguous_runs.append((run_start, run_end))

                for run_start, run_end in contiguous_runs:
                    start_col_str = f"{run_start:02d}"
                    end_col_str = f"{run_end:02d}"
                    span_cells = run_end - run_start + 1
                    beam_coordinate = f"{rack}[{start_col_str}-{end_col_str}]:{row_label}"

                    beam_segments.append(
                        {
                            "Beam_Coordinate": beam_coordinate,
                            "Rack": rack,
                            "Column_Start": start_col_str,
                            "Column_End": end_col_str,
                            "Row": row_label,
                            "Spanned_Locations": str(span_cells),
                        }
                    )

                    for col_num in range(run_start, run_end + 1):
                        beam_for_location[(rack, col_num, row_label)] = {
                            "Beam_Coordinate": beam_coordinate,
                            "Spanned_Locations": str(span_cells),
                        }

    # Copy beam assignment to alias coordinates that must behave identically.
    for target_key, source_key in BEAM_ALIASES.items():
        beam_data = beam_for_location.get(source_key)
        if beam_data is not None:
            beam_for_location[target_key] = dict(beam_data)

    # Generate location-level map with beam flags and mapped beam coordinate.
    for (rack, col_num), rack_rows in sorted(grouped.items()):
        column_str = f"{col_num:02d}"
        ordered_rows = sorted(rack_rows, key=_location_sort_key)
        for row in ordered_rows:
            row_label = str(row.get("Row", "")).strip()
            location = str(row.get("Location", "")).strip()
            beam_data = beam_for_location.get((rack, col_num, row_label), {})
            has_beam = "YES" if (beam_data and not _is_row_without_beam_support(row_label)) else "NO"

            location_map.append(
                {
                    "Location": location,
                    "Rack": rack,
                    "Column": column_str,
                    "Row": row_label,
                    "Beam_Supported": has_beam,
                    "Has_Grid": has_beam,
                    "Beam_Coordinate": beam_data.get("Beam_Coordinate", ""),
                    "Spanned_Locations": beam_data.get("Spanned_Locations", ""),
                }
            )

    grid_counts_by_beam: dict[str, int] = defaultdict(int)
    for row in location_map:
        if row.get("Beam_Supported") == "YES":
            beam_coordinate = str(row.get("Beam_Coordinate", "")).strip()
            if beam_coordinate:
                grid_counts_by_beam[beam_coordinate] += 1

    beam_height_lookup: dict[str, tuple[str, str, float, float]] = {}
    row_height_reference: dict[tuple[str, str], tuple[float, float]] = {}

    # Estimate beam height envelopes by accumulating location heights below each row.
    for beam_segment in beam_segments:
        rack = str(beam_segment.get("Rack", "")).strip()
        row_label = str(beam_segment.get("Row", "")).strip()
        start_col = _to_int(beam_segment.get("Column_Start"))
        end_col = _to_int(beam_segment.get("Column_End"))
        beam_coordinate = str(beam_segment.get("Beam_Coordinate", "")).strip()

        samples: list[tuple[float, float]] = []
        for col_num in range(start_col, end_col + 1):
            if row_index.get((rack, col_num, row_label)) is None:
                continue

            ordered_rows = sorted(grouped[(rack, col_num)], key=_location_sort_key)
            cumulative_height_cm = 0.0
            beam_count_below = 0
            for current_row in ordered_rows:
                current_row_label = str(current_row.get("Row", "")).strip()
                if current_row_label == row_label:
                    bottom_cm = cumulative_height_cm + (beam_count_below * BEAM_HEIGHT_CM)
                    top_cm = bottom_cm + BEAM_HEIGHT_CM
                    samples.append((bottom_cm, top_cm))
                    break

                current_beam_supported = bool(
                    beam_for_location.get((rack, col_num, current_row_label), {})
                ) and not _is_row_without_beam_support(current_row_label)
                cumulative_height_cm += _to_float_default(current_row.get("Location height"))
                if current_beam_supported:
                    beam_count_below += 1

        if samples:
            bottoms = [sample[0] for sample in samples]
            tops = [sample[1] for sample in samples]
            beam_bottom_cm = max(bottoms)
            beam_top_cm = max(tops)
            row_height_reference.setdefault((rack, row_label), (beam_bottom_cm, beam_top_cm))
            beam_height_lookup[beam_coordinate] = (rack, row_label, beam_bottom_cm, beam_top_cm)

    beam_height_rows: list[dict[str, str]] = []
    # Build final beam-height rows, falling back to row-level references when needed.
    for beam_segment in beam_segments:
        rack = str(beam_segment.get("Rack", "")).strip()
        row_label = str(beam_segment.get("Row", "")).strip()
        beam_coordinate = str(beam_segment.get("Beam_Coordinate", "")).strip()
        spanned_locations = str(beam_segment.get("Spanned_Locations", "")).strip()
        column_start = str(beam_segment.get("Column_Start", "")).strip()
        column_end = str(beam_segment.get("Column_End", "")).strip()

        direct_height = beam_height_lookup.get(beam_coordinate)
        if direct_height is not None:
            _, _, beam_bottom_cm, beam_top_cm = direct_height
        else:
            reference = row_height_reference.get((rack, row_label))
            if reference is None:
                continue
            beam_bottom_cm, beam_top_cm = reference

        beam_height_rows.append(
            {
                "Beam_Coordinate": beam_coordinate,
                "Rack": rack,
                "Column_Start": column_start,
                "Column_End": column_end,
                "Row": row_label,
                "Beam_Column_Count": spanned_locations,
                "Grid_Count": str(grid_counts_by_beam.get(beam_coordinate, 0)),
                "Beam_Bottom_cm": f"{beam_bottom_cm:.2f}",
                "Beam_Top_cm": f"{beam_top_cm:.2f}",
                "Beam_Height_Range_cm": f"{beam_bottom_cm:.0f}-{beam_top_cm:.0f}",
            }
        )

    beam_range_by_coordinate = {
        str(row.get("Beam_Coordinate", "")).strip(): str(row.get("Beam_Height_Range_cm", "")).strip()
        for row in beam_height_rows
        if str(row.get("Beam_Coordinate", "")).strip()
    }
    for row in location_map:
        beam_coordinate = str(row.get("Beam_Coordinate", "")).strip()
        row["Beam_Height_Range_cm"] = beam_range_by_coordinate.get(beam_coordinate, "")

    location_file = BEAM_OUTPUT_DIR / "Location_Beam_Map.csv"
    beam_height_file = BEAM_OUTPUT_DIR / "Beam_Height_Coordinates.csv"

    with location_file.open("w", newline="", encoding="utf-8") as target:
        fieldnames = [
            "Location",
            "Rack",
            "Column",
            "Row",
            "Beam_Supported",
            "Has_Grid",
            "Beam_Coordinate",
            "Beam_Height_Range_cm",
            "Spanned_Locations",
        ]
        writer = csv.DictWriter(target, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(location_map)

    with beam_height_file.open("w", newline="", encoding="utf-8") as target:
        fieldnames = [
            "Beam_Coordinate",
            "Rack",
            "Column_Start",
            "Column_End",
            "Row",
            "Beam_Column_Count",
            "Grid_Count",
            "Beam_Bottom_cm",
            "Beam_Top_cm",
            "Beam_Height_Range_cm",
        ]
        writer = csv.DictWriter(target, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(beam_height_rows)

    return BEAM_OUTPUT_DIR


if __name__ == "__main__":
    # Stage 01 entrypoint: prepare location table, then derive beam/grid mappings.
    output_path = prepare_location_data()
    utilization_path = calculate_current_layout_space_utilization()
    beam_output_path = build_beam_grid_map()
    print(f"Data preparation complete. Output written to: {output_path}")
    print(f"Current layout space utilization written to: {utilization_path}")
    print(f"Beam mapping complete. Output written to: {beam_output_path}")
